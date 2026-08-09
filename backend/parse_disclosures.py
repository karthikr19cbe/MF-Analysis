"""
Main parser for Indian mutual fund SEBI monthly portfolio disclosure XLSX files.
Reads all XLSX files from ../disclosures/, extracts holdings, deduplicates by ISIN,
computes aggregations, and writes ../data/consolidated.json.
"""

import json
import logging
import os
import shutil
import sys
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

from normalizer import (
    normalize_name,
    normalize_scheme_name,
    detect_cash_equivalent,
    find_header_row,
    detect_unit_multiplier,
    extract_scheme_name,
    extract_amc_name,
)
from hhi import calculate_hhi, interpret_hhi

# Paths
BASE_DIR = Path(__file__).resolve().parent.parent
DISCLOSURES_DIR = BASE_DIR / "disclosures"
DATA_DIR = BASE_DIR / "data"
OUTPUT_FILE = DATA_DIR / "consolidated.json"
ERROR_LOG = DATA_DIR / "parse_errors.log"

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger(__name__)


def parse_sheet(sheet, filename: str, amc: str) -> list[dict]:
    """Parse a single worksheet and return a list of holding records."""
    header_row, col_map = find_header_row(sheet)

    if header_row < 0 or "stock_name" not in col_map:
        logger.warning(f"  Skipping sheet '{sheet.title}' in {filename}: no valid header found")
        return []

    # If no sector column but rating column exists, check if rating contains sector data
    if "sector" not in col_map and "rating" in col_map:
        # Sample a few data rows to see if the "rating" column has sector-like values
        credit_rating_patterns = {"aaa", "aa+", "aa-", "aa", "a+", "a-", "a1+", "a1", "bbb", "bb", "b", "crisil", "icra", "care", "fitch", "sovereign"}
        is_sector = True
        for sample_row in range(header_row + 1, min(header_row + 6, sheet.max_row + 1)):
            val = sheet.cell(row=sample_row, column=col_map["rating"]).value
            if val and isinstance(val, str):
                val_lower = val.strip().lower()
                if val_lower in credit_rating_patterns or len(val_lower) <= 4:
                    is_sector = False
                    break
        if is_sector:
            col_map["sector"] = col_map["rating"]
            logger.info(f"  Using 'Rating' column as sector for sheet '{sheet.title}'")

    scheme_name = normalize_scheme_name(extract_scheme_name(sheet, header_row, col_map))

    # Detect value units
    multiplier = 1.0
    if "market_value" in col_map:
        multiplier = detect_unit_multiplier(sheet, header_row, col_map["market_value"])

    # Section header patterns -> asset_class mapping
    SECTION_PATTERNS = [
        ("equity & equity related", "Equity"),
        ("equity &amp; equity", "Equity"),
        ("(a) listed", None),  # sub-section, keep current class
        ("(b) reit", "REIT"),
        ("(b) invit", "REIT"),
        ("reits & invit", "REIT"),
        ("reit", "REIT"),
        ("invit", "REIT"),
        ("(b) unlisted", None),
        ("(c) ", None),
        ("debt instrument", "Debt"),
        ("money market instrument", "Money Market"),
        ("certificate of deposit", "Money Market"),
        ("commercial paper", "Commercial Paper"),
        ("treasury bill", "Treasury Bills"),
        ("exchange traded fund", "Commodities"),  # gold/silver ETFs in multi-asset funds
        ("gold etf", "Commodities"),
        ("silver etf", "Commodities"),
        ("mutual fund unit", "Mutual Fund Units"),
        ("reverse repo", "TREPS"),
        ("treps", "TREPS"),
        ("tri-party repo", "TREPS"),
        ("derivative", "Derivatives"),
        ("index / stock future", "Derivatives"),
        ("index / stock option", "Derivatives"),
        ("future", "Derivatives"),
        ("option", "Derivatives"),
        ("hedging position", "Derivatives"),
        ("securitised", "Debt"),
        ("fixed deposit", "Debt"),
        ("net receivable", "Net Receivables/Payables"),
        ("net payable", "Net Receivables/Payables"),
        ("net current asset", "Net Receivables/Payables"),
    ]

    # Rows that are purely section headers / metadata (skip as data)
    SKIP_PATTERNS = [
        "total", "grand total", "sub total", "subtotal", "sub-total",
        "(a) listed", "(b) unlisted", "(c) ",
        "net assets value", "net asset value",
        "portfolio turnover", "investment in repo",
        "total exposure", "total outstanding",
        "total value and percentage", "total number of contract",
        "gross notional", "not applicable",
        "primary investment objective",
        "debt instruments having structured",
        "for the period ended",
    ]

    # Asset classes that are SUB-sections nested under a main section (e.g. a
    # "(b) Reits" block inside "Equity & Equity related"). They apply only until
    # their own "Sub Total" row, after which we revert to the enclosing main
    # section's class. This prevents the sticky class from leaking onto sibling
    # rows listed after the sub-block (e.g. equity shares listed after a REIT
    # sub-total — as happens with the Vedanta demerger entities).
    # REIT/InvIT nest under Equity; Treasury Bills nest under Money Market.
    # Both revert to their enclosing main section on the closing "Sub Total".
    SUBSECTION_CLASSES = {"REIT", "Treasury Bills"}

    holdings = []
    current_main_class = "Equity"   # Enclosing main section (revert target)
    current_asset_class = "Equity"  # Effective class for the current row

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        name_val = sheet.cell(row=row_idx, column=col_map["stock_name"]).value
        if not name_val or not isinstance(name_val, str) or len(name_val.strip()) < 2:
            continue

        name_val = name_val.strip()
        name_lower = name_val.lower()

        # Stop at the portfolio GRAND TOTAL — anything after it is supplementary
        # disclosure (derivative/commodity-future detail tables, notes, symbol
        # legends) that uses a different column layout and would double-count
        # exposure, pushing the portfolio weight above 100%.
        if "grand total" in name_lower:
            break

        # Check if this row is a section header that changes the asset class
        # Only treat as header if the row has no numeric market value (real data rows have values)
        has_market_value = False
        if "market_value" in col_map:
            mv_check = sheet.cell(row=row_idx, column=col_map["market_value"]).value
            if mv_check is not None:
                try:
                    mv_num = float(str(mv_check).replace(",", ""))
                    if mv_num != 0:
                        has_market_value = True
                except (ValueError, TypeError):
                    pass

        is_section_header = False
        if not has_market_value:
            for pattern, asset_class in SECTION_PATTERNS:
                if pattern in name_lower:
                    if asset_class is not None:
                        current_asset_class = asset_class
                        # Main sections also become the revert target; sub-section
                        # classes (REIT) only override the effective class.
                        if asset_class not in SUBSECTION_CLASSES:
                            current_main_class = asset_class
                    is_section_header = True
                    break

        if is_section_header:
            continue

        # A "Sub Total" closes the current (sub-)section. Revert the effective
        # class to the enclosing main section so following sibling rows aren't
        # mislabeled with a leaked sub-section class (e.g. REIT -> Equity).
        if "sub total" in name_lower or "subtotal" in name_lower or "sub-total" in name_lower:
            current_asset_class = current_main_class
            continue

        # Skip summary/metadata rows
        if any(skip in name_lower for skip in SKIP_PATTERNS):
            continue

        isin = None
        if "isin" in col_map:
            isin_val = sheet.cell(row=row_idx, column=col_map["isin"]).value
            if isin_val and isinstance(isin_val, str) and len(isin_val.strip()) >= 10:
                isin = isin_val.strip().upper()

        sector = ""
        if "sector" in col_map:
            sec_val = sheet.cell(row=row_idx, column=col_map["sector"]).value
            if sec_val and isinstance(sec_val, str):
                sector = sec_val.strip()

        quantity = 0
        if "quantity" in col_map:
            qty_val = sheet.cell(row=row_idx, column=col_map["quantity"]).value
            if qty_val is not None:
                try:
                    quantity = int(float(str(qty_val).replace(",", "")))
                except (ValueError, TypeError):
                    quantity = 0

        market_value = 0.0
        if "market_value" in col_map:
            mv_val = sheet.cell(row=row_idx, column=col_map["market_value"]).value
            if mv_val is not None:
                try:
                    mv_str = str(mv_val).replace(",", "").strip()
                    # Handle parenthesized negatives: (1791.29) -> -1791.29
                    if mv_str.startswith("(") and mv_str.endswith(")"):
                        mv_str = "-" + mv_str[1:-1]
                    market_value = float(mv_str) * multiplier
                except (ValueError, TypeError):
                    market_value = 0.0

        pct_net_assets = 0.0
        if "pct_net_assets" in col_map:
            pct_val = sheet.cell(row=row_idx, column=col_map["pct_net_assets"]).value
            if pct_val is not None:
                try:
                    pct_str = str(pct_val).replace(",", "").replace("%", "").strip()
                    # Handle parenthesized negatives: (6.78) -> -6.78
                    if pct_str.startswith("(") and pct_str.endswith(")"):
                        pct_str = "-" + pct_str[1:-1]
                    pct_net_assets = float(pct_str)
                except (ValueError, TypeError):
                    pct_net_assets = 0.0

        # Detect cash equivalents
        cash_category = detect_cash_equivalent(name_val, isin)

        # Override asset_class from cash detection if applicable
        asset_class = current_asset_class
        if cash_category:
            cash_to_class = {
                "Cash": "Net Receivables/Payables",
                "TREPS": "TREPS",
                "CBLO": "TREPS",
                "Reverse Repo": "TREPS",
                "Treasury Bills": "Treasury Bills",
                "Commercial Paper": "Commercial Paper",
                "Certificate of Deposit": "Money Market",
                "Fixed Deposit": "Debt",
                "Net Receivables": "Net Receivables/Payables",
                "Net Current Assets": "Net Receivables/Payables",
                "Government Securities": "Debt",
            }
            asset_class = cash_to_class.get(cash_category, asset_class)

        # Commodity ETFs (gold/silver) in multi-asset funds — classify by name so
        # they are never mislabeled by a leaked section class (e.g. Treasury Bills).
        if ("gold" in name_lower or "silver" in name_lower) and ("etf" in name_lower or "bees" in name_lower):
            asset_class = "Commodities"

        # Skip rows with no value and no quantity (metadata/footnote rows)
        if market_value == 0 and quantity == 0 and pct_net_assets == 0:
            continue

        holdings.append({
            "raw_name": name_val,
            "name": normalize_name(name_val),
            "isin": isin,
            "sector": sector,
            "quantity": quantity,
            "market_value_lakhs": round(market_value, 4),
            "pct_of_net_assets": round(pct_net_assets, 4),
            "scheme_name": scheme_name,
            "amc": amc,
            "cash_category": cash_category,
            "asset_class": asset_class,
        })

    # Auto-detect if percentages are stored as decimals (sum ≈ 1.0 instead of ≈ 100)
    if holdings:
        total_pct = sum(abs(h["pct_of_net_assets"]) for h in holdings)
        if 0.5 < total_pct < 2.0 and len(holdings) > 3:
            for h in holdings:
                h["pct_of_net_assets"] = round(h["pct_of_net_assets"] * 100, 4)

    return holdings


class _XlrdSheetWrapper:
    """Wraps an xlrd sheet to provide an openpyxl-compatible interface."""

    def __init__(self, xlrd_sheet):
        self._sheet = xlrd_sheet
        self.title = xlrd_sheet.name
        self.max_row = xlrd_sheet.nrows
        self.max_column = xlrd_sheet.ncols
        self.merged_cells = _MergedCellsWrapper(xlrd_sheet.merged_cells)

    def cell(self, row, column):
        """Return a cell-like object (1-indexed row/col like openpyxl)."""
        try:
            value = self._sheet.cell_value(row - 1, column - 1)
            if value == "":
                value = None
        except (IndexError, Exception):
            value = None
        return _CellWrapper(value)


class _CellWrapper:
    def __init__(self, value):
        self.value = value


class _MergedCellsWrapper:
    """Wraps xlrd merged_cells to provide openpyxl-compatible ranges."""

    def __init__(self, merged_cells):
        self._ranges = []
        for rlo, rhi, clo, chi in merged_cells:
            self._ranges.append(_MergedRange(rlo + 1, rhi, clo + 1, chi))

    @property
    def ranges(self):
        return self._ranges


class _MergedRange:
    def __init__(self, min_row, max_row, min_col, max_col):
        self.min_row = min_row
        self.max_row = max_row
        self.min_col = min_col
        self.max_col = max_col


class _XlrdWorkbookWrapper:
    """Wraps an xlrd workbook to provide an openpyxl-compatible interface."""

    def __init__(self, xlrd_wb):
        self._wb = xlrd_wb
        self.sheetnames = xlrd_wb.sheet_names()

    def __getitem__(self, name):
        return _XlrdSheetWrapper(self._wb.sheet_by_name(name))

    def close(self):
        self._wb.release_resources()


def _open_workbook(filepath: Path):
    """Open a workbook, handling both .xlsx and legacy .xls files."""
    # Try openpyxl first (handles .xlsx and some .xls-renamed-xlsx)
    try:
        return openpyxl.load_workbook(filepath, read_only=False, data_only=True), None
    except Exception:
        pass

    # For .xls files: try xlrd first (handles genuine legacy .xls), then renamed-xlsx fallback
    if filepath.suffix.lower() == ".xls":
        # Try as legacy .xls via xlrd first
        try:
            import xlrd
            xlrd_wb = xlrd.open_workbook(str(filepath))
            return _XlrdWorkbookWrapper(xlrd_wb), None
        except ImportError:
            pass
        except Exception:
            pass

        # Fallback: try as renamed xlsx
        tmp_path = Path(tempfile.mkdtemp()) / (filepath.stem + "_tmp.xlsx")
        shutil.copy2(filepath, tmp_path)
        try:
            wb = openpyxl.load_workbook(tmp_path, read_only=False, data_only=True)
            return wb, tmp_path
        except Exception:
            tmp_path.unlink(missing_ok=True)

    raise ValueError(f"Cannot open {filepath.name}")


# Sheet names to skip (index/TOC sheets)
SKIP_SHEETS = {"index", "toc", "table of contents", "summary", "sheet1", "abali"}


def parse_file(filepath: Path) -> list[dict]:
    """Parse all sheets in an XLSX file."""
    filename = filepath.name
    amc = extract_amc_name(filename)
    logger.info(f"Parsing: {filename} (AMC: {amc})")

    tmp_path = None
    try:
        wb, tmp_path = _open_workbook(filepath)
    except Exception as e:
        logger.error(f"  Failed to open {filename}: {e}")
        return []

    all_holdings = []
    try:
        for sheet_name in wb.sheetnames:
            # Skip index/TOC sheets
            if sheet_name.lower().strip() in SKIP_SHEETS:
                logger.info(f"  Skipping index sheet: '{sheet_name}'")
                continue
            sheet = wb[sheet_name]
            if sheet.max_row is None or sheet.max_row < 3:
                continue
            holdings = parse_sheet(sheet, filename, amc)
            if holdings:
                logger.info(f"  Sheet '{sheet_name}': {len(holdings)} holdings extracted")
                all_holdings.extend(holdings)
    finally:
        wb.close()
        if tmp_path:
            tmp_path.unlink(missing_ok=True)

    return all_holdings


def detect_date_from_filename(filename: str) -> str:
    """Try to extract a date from the filename."""
    import re
    # Try YYYY-MM-DD or YYYY_MM_DD
    match = re.search(r"(\d{4})[-_](\d{2})[-_](\d{2})", filename)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"

    # Try month name patterns
    months = {
        "jan": "01", "feb": "02", "mar": "03", "apr": "04",
        "may": "05", "jun": "06", "jul": "07", "aug": "08",
        "sep": "09", "oct": "10", "nov": "11", "dec": "12",
    }
    for mon, num in months.items():
        pattern = rf"({mon})\w*[-_]?(\d{{2,4}})"
        match = re.search(pattern, filename, re.IGNORECASE)
        if match:
            year = match.group(2)
            if len(year) == 2:
                year = "20" + year
            return f"{year}-{num}-01"

    return datetime.now().strftime("%Y-%m-%d")


def aggregate(all_holdings: list[dict], files_parsed: list[dict]) -> dict:
    """Aggregate all holdings into the consolidated data structure."""

    # --- Group by ISIN (primary) or normalized name (fallback) ---
    stocks_by_key = defaultdict(lambda: {
        "isin": None,
        "name": "",
        "name_variants": set(),
        "sector": "",
        "asset_class": "Equity",
        "asset_class_votes": defaultdict(int),
        "sector_votes": defaultdict(int),
        "total_market_value_lakhs": 0.0,
        "total_pct_sum": 0.0,
        "fund_count": 0,
        "funds": [],
        "schemes_seen": set(),
    })

    # Fund-level data
    funds_data = defaultdict(lambda: {
        "amc": "",
        "date": "",
        "total_aum_lakhs": 0.0,
        "holding_count": 0,
        "cash_and_equivalents": {
            "cash": 0.0,
            "reverse_repo": 0.0,
            "treps": 0.0,
            "treasury_bills": 0.0,
            "government_securities": 0.0,
            "fixed_deposit": 0.0,
            "certificate_of_deposit": 0.0,
            "commercial_paper": 0.0,
            "net_receivables": 0.0,
            "net_current_assets": 0.0,
            "other": 0.0,
            "total_lakhs": 0.0,
            "pct_of_net_assets": 0.0,
        },
    })

    for h in all_holdings:
        scheme = h["scheme_name"]
        fund = funds_data[scheme]
        fund["amc"] = h["amc"]
        fund["total_aum_lakhs"] += h["market_value_lakhs"]
        fund["holding_count"] += 1

        # Track cash equivalents for liquidity summary
        if h["cash_category"]:
            cat = h["cash_category"]
            cash = fund["cash_and_equivalents"]
            cat_key = cat.lower().replace(" ", "_")
            if cat_key in cash:
                cash[cat_key] += h["market_value_lakhs"]
            else:
                cash["other"] += h["market_value_lakhs"]
            cash["total_lakhs"] += h["market_value_lakhs"]
            cash["pct_of_net_assets"] += h["pct_of_net_assets"]

        # Use ISIN as primary key, fall back to normalized name
        if h["isin"]:
            key = h["isin"]
        else:
            key = f"NAME__{h['name']}"

        stock = stocks_by_key[key]
        if h["isin"]:
            stock["isin"] = h["isin"]
        stock["name_variants"].add(h["raw_name"])

        # Use the most common/clean name
        if not stock["name"] or len(h["name"]) > len(stock["name"]):
            stock["name"] = h["name"]

        # Vote on sector and asset class
        if h["sector"]:
            stock["sector_votes"][h["sector"]] += 1
        stock["asset_class_votes"][h["asset_class"]] += 1

        stock["total_market_value_lakhs"] += h["market_value_lakhs"]
        stock["total_pct_sum"] += h["pct_of_net_assets"]

        if scheme not in stock["schemes_seen"]:
            stock["schemes_seen"].add(scheme)
            stock["fund_count"] += 1

        stock["funds"].append({
            "scheme_name": scheme,
            "amc": h["amc"],
            "date": detect_date_from_filename(h["amc"]),
            "quantity": h["quantity"],
            "market_value_lakhs": h["market_value_lakhs"],
            "pct_of_net_assets": h["pct_of_net_assets"],
        })

    # --- Resolve sectors and asset classes (most frequent vote per stock) ---
    for key, stock in stocks_by_key.items():
        if stock["sector_votes"]:
            stock["sector"] = max(stock["sector_votes"], key=stock["sector_votes"].get)
        if stock["asset_class_votes"]:
            # Deterministic tie-break: most votes wins; on a tie prefer "Equity"
            # (the safe default for a security with a real industry sector),
            # otherwise fall back to alphabetical order. Avoids non-deterministic
            # dict-ordering flips that would split a holding across periods.
            votes = stock["asset_class_votes"]
            best = max(votes.values())
            tied = sorted(ac for ac, n in votes.items() if n == best)
            stock["asset_class"] = "Equity" if "Equity" in tied else tied[0]

    # --- Build output structures ---
    total_schemes = len(funds_data)
    total_aum = sum(f["total_aum_lakhs"] for f in funds_data.values())

    # Stocks output
    stocks_output = {}
    stock_weights = []
    for key, stock in stocks_by_key.items():
        weight = stock["total_market_value_lakhs"] / total_aum if total_aum > 0 else 0
        stock_weights.append(weight)

        weighted_avg = (stock["total_market_value_lakhs"] / total_aum * 100) if total_aum > 0 else 0

        stocks_output[key] = {
            "isin": stock["isin"],
            "name": stock["name"],
            "name_variants": sorted(stock["name_variants"]),
            "sector": stock["sector"],
            "asset_class": stock["asset_class"],
            "total_market_value_lakhs": round(stock["total_market_value_lakhs"], 2),
            "weighted_avg_pct": round(weighted_avg, 2),
            "fund_count": stock["fund_count"],
            "funds": stock["funds"],
        }

    # Sectors output
    sectors = defaultdict(lambda: {"total_market_value_lakhs": 0.0, "stock_count": 0, "weighted_avg_pct": 0.0})
    sector_isins = defaultdict(set)
    for key, stock in stocks_output.items():
        sec = stock["sector"] or "Unclassified"
        sectors[sec]["total_market_value_lakhs"] += stock["total_market_value_lakhs"]
        sector_isins[sec].add(key)

    # Asset classes summary
    asset_classes = defaultdict(lambda: {"total_market_value_lakhs": 0.0, "holding_count": 0, "weighted_avg_pct": 0.0})
    for key, stock in stocks_output.items():
        ac = stock["asset_class"]
        asset_classes[ac]["total_market_value_lakhs"] += stock["total_market_value_lakhs"]
        asset_classes[ac]["holding_count"] += 1

    for ac in asset_classes:
        asset_classes[ac]["total_market_value_lakhs"] = round(asset_classes[ac]["total_market_value_lakhs"], 2)
        asset_classes[ac]["weighted_avg_pct"] = round(
            (asset_classes[ac]["total_market_value_lakhs"] / total_aum * 100) if total_aum > 0 else 0, 2
        )

    for sec in sectors:
        sectors[sec]["stock_count"] = len(sector_isins[sec])
        sectors[sec]["total_market_value_lakhs"] = round(sectors[sec]["total_market_value_lakhs"], 2)
        sectors[sec]["weighted_avg_pct"] = round(
            (sectors[sec]["total_market_value_lakhs"] / total_aum * 100) if total_aum > 0 else 0, 2
        )

    # Sector-level HHI
    sector_weights = [s["total_market_value_lakhs"] / total_aum for s in sectors.values()] if total_aum > 0 else []

    # Funds output
    funds_output = {}
    for scheme, fund in funds_data.items():
        funds_output[scheme] = {
            "amc": fund["amc"],
            "date": detect_date_from_filename(fund["amc"]),
            "total_aum_lakhs": round(fund["total_aum_lakhs"], 2),
            "holding_count": fund["holding_count"],
            "cash_and_equivalents": {
                k: round(v, 2) if isinstance(v, float) else v
                for k, v in fund["cash_and_equivalents"].items()
            },
        }

    # HHI
    stock_hhi = calculate_hhi(stock_weights)
    sector_hhi = calculate_hhi(sector_weights)

    # Top 10 weight
    sorted_stocks = sorted(stocks_output.values(), key=lambda s: s["total_market_value_lakhs"], reverse=True)
    top_10_value = sum(s["total_market_value_lakhs"] for s in sorted_stocks[:10])
    top_10_weight = (top_10_value / total_aum * 100) if total_aum > 0 else 0

    # High conviction stocks (appearing in most funds)
    high_conviction = sorted(
        [s for s in stocks_output.values() if s["fund_count"] >= 2],
        key=lambda s: (-s["fund_count"], -s["weighted_avg_pct"]),
    )[:20]

    high_conviction_output = [
        {
            "isin": s["isin"],
            "name": s["name"],
            "fund_count": s["fund_count"],
            "total_funds": total_schemes,
            "appearance_pct": round(s["fund_count"] / total_schemes * 100, 1) if total_schemes > 0 else 0,
            "avg_weight_pct": s["weighted_avg_pct"],
            "total_market_value_lakhs": s["total_market_value_lakhs"],
            "sector": s["sector"],
            "asset_class": s.get("asset_class", "Equity"),
        }
        for s in high_conviction
    ]

    # Liquidity summary
    total_cash = sum(f["cash_and_equivalents"]["total_lakhs"] for f in funds_data.values())

    return {
        "meta": {
            "generated_at": datetime.now().isoformat(),
            "file_count": len(files_parsed),
            "files_parsed": files_parsed,
            "total_schemes": total_schemes,
            "total_unique_stocks": len(stocks_output),
        },
        "stocks": stocks_output,
        "sectors": dict(sectors),
        "asset_classes": dict(asset_classes),
        "funds": funds_output,
        "concentration": {
            "hhi": {
                "by_stock": round(stock_hhi, 4),
                "by_sector": round(sector_hhi, 4),
                "interpretation": interpret_hhi(stock_hhi),
            },
            "top_10_weight_pct": round(top_10_weight, 2),
            "high_conviction": high_conviction_output,
        },
        "liquidity_summary": {
            "total_cash_lakhs": round(total_cash, 2),
            "total_aum_lakhs": round(total_aum, 2),
            "cash_pct": round((total_cash / total_aum * 100) if total_aum > 0 else 0, 2),
        },
    }


def main():
    """Main entry point."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # Find all XLSX files
    xlsx_files = list(DISCLOSURES_DIR.glob("*.xlsx")) + list(DISCLOSURES_DIR.glob("*.xls"))
    if not xlsx_files:
        logger.warning(f"No XLSX files found in {DISCLOSURES_DIR}")
        logger.info("Please add mutual fund disclosure XLSX files to the disclosures/ folder.")
        # Write empty structure so frontend doesn't break
        empty = {
            "meta": {
                "generated_at": datetime.now().isoformat(),
                "file_count": 0,
                "files_parsed": [],
                "total_schemes": 0,
                "total_unique_stocks": 0,
            },
            "stocks": {},
            "sectors": {},
            "funds": {},
            "concentration": {
                "hhi": {"by_stock": 0, "by_sector": 0, "interpretation": "N/A"},
                "top_10_weight_pct": 0,
                "high_conviction": [],
            },
            "liquidity_summary": {"total_cash_lakhs": 0, "total_aum_lakhs": 0, "cash_pct": 0},
        }
        with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
            json.dump(empty, f, indent=2, ensure_ascii=False)
        logger.info(f"Empty consolidated.json written to {OUTPUT_FILE}")
        return

    # Parse all files
    all_holdings = []
    files_parsed = []
    errors = []

    for filepath in xlsx_files:
        try:
            holdings = parse_file(filepath)
            schemes = set(h["scheme_name"] for h in holdings)
            files_parsed.append({
                "filename": filepath.name,
                "scheme_count": len(schemes),
                "date_detected": detect_date_from_filename(filepath.name),
                "holdings_count": len(holdings),
            })
            all_holdings.extend(holdings)
        except Exception as e:
            error_msg = f"Error parsing {filepath.name}: {e}"
            logger.error(error_msg)
            errors.append(error_msg)

    logger.info(f"\nTotal: {len(all_holdings)} holdings from {len(files_parsed)} files")

    # Write error log
    if errors:
        with open(ERROR_LOG, "w", encoding="utf-8") as f:
            f.write("\n".join(errors))
        logger.warning(f"Errors logged to {ERROR_LOG}")

    # Aggregate and write output
    result = aggregate(all_holdings, files_parsed)

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    logger.info(f"Consolidated data written to {OUTPUT_FILE}")
    logger.info(f"  Unique stocks: {result['meta']['total_unique_stocks']}")
    logger.info(f"  Schemes: {result['meta']['total_schemes']}")
    logger.info(f"  Stock HHI: {result['concentration']['hhi']['by_stock']} ({result['concentration']['hhi']['interpretation']})")
    logger.info(f"  Sector HHI: {result['concentration']['hhi']['by_sector']}")
    logger.info(f"  Top 10 weight: {result['concentration']['top_10_weight_pct']}%")
    logger.info(f"  Cash %: {result['liquidity_summary']['cash_pct']}%")


if __name__ == "__main__":
    main()
